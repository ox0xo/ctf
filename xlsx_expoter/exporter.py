from openpyxl import load_workbook, Workbook
import os, glob, sys


def read_data(in_path, result, user_name=None):
    #wb = load_workbook(os.path.dirname(__file__) + in_path)
    wb = load_workbook(in_path)
    for name in wb.sheetnames:
        if user_name == None:
            user_name = name
        if not name in IGNORE:
            for row in wb[name]["B3":"G22"]:
                category = row[0].value
                for i in range(1, 6):
                    if row[i].value != UNDEFINED:
                        result.append([user_name, category, row[i].value, 6-i])
    return result


def write_data(result, out_path):
    wb = Workbook()
    sheet = wb.active
    sheet.title = "marged"
    cells = sheet["A1":"D%d"%(len(result)+1)]
    header = cells[0]
    header[0].value = "Person"
    header[1].value = "Category"
    header[2].value = "Skill"
    header[3].value = "Level"
    for i in range(1, len(cells)):
        cells[i][0].value = result[i-1][0]
        cells[i][1].value = result[i-1][1]
        cells[i][2].value = result[i-1][2]
        cells[i][3].value = result[i-1][3]
    wb.save(out_path)

if __name__ == "__main__":
    UNDEFINED = "（未選択）"
    IGNORE = ["業務スキル", "選択肢(知識・技術)", "選択肢(業務)"]
    base_dir = os.path.dirname(__file__) + r"/excel/**/*.xlsx"
    out_path = os.path.dirname(__file__) + r"/excel/result.xlsx"
    result = []
    for path in glob.glob(base_dir, recursive=True):
        name = path.split("_")[1].split(".")[0]
        result = read_data(path, result, name)
    write_data(result, out_path)

